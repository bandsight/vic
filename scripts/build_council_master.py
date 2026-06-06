from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import datetime, timezone
from html.parser import HTMLParser
import json
from pathlib import Path
import re
from typing import Any
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
REFERENCE_DIR = ROOT / "data" / "reference"
SOURCE_DIR = REFERENCE_DIR / "source"

GEOGRAPHY_PATH = REFERENCE_DIR / "victorian-council-geography.json"
ABS_ASGS_PATH = SOURCE_DIR / "abs-asgs2025-vic-lga-attributes.json"
LGPRF_PATH = SOURCE_DIR / "LGPRF-2020-2025-Full-Council-Data-Set-Nov25-Final-Release.xlsx"
GOVERNANCE_PATH = SOURCE_DIR / "G-and-M-Checklist-Data-Cube-All-Years-2025-Nov-25.xlsx"
VGCCC_PATH = SOURCE_DIR / "yearly_density_statistical_release_nov_24.xlsx"
VEC_CACHE_PATH = SOURCE_DIR / "vec-local-council-electoral-structure.json"

MASTER_JSON_PATH = REFERENCE_DIR / "victorian-council-master.json"
MASTER_CSV_PATH = REFERENCE_DIR / "victorian-council-master.csv"
INDICATORS_CSV_PATH = REFERENCE_DIR / "victorian-council-lgprf-indicators-2024-25.csv"
GOVERNANCE_CSV_PATH = REFERENCE_DIR / "victorian-council-governance-checklist-2024-25.csv"
STATIC_MASTER_JSON_PATH = ROOT / "static" / "data" / "victorian-council-master.json"

LGPRF_SOURCE_URL = "https://www.localgovernment.vic.gov.au/strengthening-councils/performance-reporting"
KYC_SOURCE_URL = "https://www.vic.gov.au/know-your-council-comparison-dashboard"
DATAVIC_LGPRF_URL = "https://discover.data.vic.gov.au/dataset/local-government-performance-reporting"
ABS_ASGS_SOURCE_URL = "https://geo.abs.gov.au/arcgis/rest/services/ASGS2025/LGA/MapServer/layers"
VICMAP_SOURCE_URL = "https://discover.data.vic.gov.au/dataset/vicmap-reference-administrative-local-government-area-lga-table"
VIF_SOURCE_URL = "https://www.planning.vic.gov.au/guides-and-resources/Data-spatial-and-insights/discover-and-access-planning-open-data/victoria-in-future"
VGCCC_SOURCE_URL = "https://discover.data.vic.gov.au/dataset/current-lga-population-density-gaming-expenditures-statistics"
VEC_SOURCE_URL = "https://www.vec.vic.gov.au/electoral-boundaries/local-councils"

_KEY_RE = re.compile(r"[^A-Z0-9]+")
_COUNCIL_TERMS_RE = re.compile(r"\b(RURAL CITY|CITY|SHIRE|BOROUGH|COUNCIL)\b")

MASTER_LGPRF_FIELDS = {
    "C1": "lgprf_expenses_per_head",
    "C2": "lgprf_infrastructure_per_head",
    "C3": "lgprf_population_density_per_road_length",
    "C4": "lgprf_own_source_revenue_per_head",
    "C5": "lgprf_recurrent_grants_per_head",
    "C6": "lgprf_relative_socioeconomic_disadvantage",
    "C7": "lgprf_staff_turnover_pct",
    "E4": "lgprf_average_rate_per_property_assessment",
    "G2": "lgprf_satisfaction_consultation",
    "G3": "lgprf_councillor_attendance_pct",
    "G4": "lgprf_cost_elected_representation",
    "G5": "lgprf_satisfaction_council_decisions",
    "OP1": "lgprf_adjusted_underlying_result_pct",
    "O5": "lgprf_asset_renewal_upgrade_to_depreciation_pct",
    "R5": "lgprf_satisfaction_sealed_local_roads",
    "S1": "lgprf_rates_to_adjusted_underlying_revenue_pct",
    "S2": "lgprf_rates_to_property_values_pct",
}

MASTER_COLUMNS = [
    "council_key",
    "short_name",
    "long_name",
    "status",
    "is_active",
    "council_category",
    "council_type",
    "official_name",
    "spatial_name",
    "spatial_key",
    "map_join_key",
    "lga_code",
    "abs_lga_code",
    "abs_lga_code_2025",
    "abs_lga_name_2025",
    "abs_area_albers_sqkm",
    "abs_shape_area",
    "abs_shape_length",
    "abs_join_method",
    "state",
    "polygon_record_count",
    "office_township",
    "office_address",
    "office_lat",
    "office_lon",
    "office_geocoded",
    "vif_metropolitan_region",
    "vif_regional_partnership",
    "vif_region_note",
    "vec_electoral_structure",
    "vec_ward_count",
    "vec_councillor_count",
    "vec_source_url",
    "lgprf_latest_year",
    "lgprf_group",
    *MASTER_LGPRF_FIELDS.values(),
    "governance_latest_year",
    "governance_item_count",
    "governance_yes_count",
    "governance_no_count",
    "governance_yes_pct",
    "vgccc_year",
    "vgccc_region",
    "vgccc_lga_code",
    "vgccc_net_expenditure",
    "vgccc_seifa_dis_score",
    "vgccc_seifa_dis_rank_state",
    "vgccc_seifa_dis_rank_country",
    "vgccc_seifa_dis_rank_metro",
    "vgccc_seifa_advdis_score",
    "vgccc_seifa_advdis_rank_state",
    "vgccc_seifa_advdis_rank_country",
    "vgccc_seifa_advdis_rank_metro",
    "vgccc_adult_population_2024",
    "vgccc_adults_per_venue_2024",
    "vgccc_egms_per_1000_adults_2024",
    "vgccc_expenditure_per_adult_2024",
    "vgccc_workforce_june_2024",
    "vgccc_unemployed_june_2024",
    "vgccc_unemployment_rate_june_2024",
    "has_abs_asgs",
    "has_lgprf",
    "has_governance",
    "has_vec",
    "has_vgccc",
]

VIF_METROPOLITAN_REGIONS = {
    "Inner Metropolitan": ["Melbourne", "Port Phillip", "Yarra"],
    "Inner South East": ["Bayside", "Boroondara", "Glen Eira", "Stonnington"],
    "Western": ["Brimbank", "Hobsons Bay", "Maribyrnong", "Melton", "Moonee Valley", "Wyndham"],
    "Northern": ["Banyule", "Hume", "Darebin", "Merri-bek", "Nillumbik", "Whittlesea"],
    "Northern (part)": ["Mitchell"],
    "Eastern": ["Knox", "Manningham", "Maroondah", "Monash", "Whitehorse", "Yarra Ranges"],
    "Southern": ["Cardinia", "Casey", "Frankston", "Greater Dandenong", "Kingston", "Mornington Peninsula"],
}

VIF_REGIONAL_PARTNERSHIPS = {
    "Barwon": ["Colac Otway", "Greater Geelong", "Queenscliffe", "Surf Coast"],
    "Central Highlands": ["Ararat", "Ballarat", "Golden Plains", "Hepburn", "Moorabool", "Pyrenees"],
    "Gippsland": ["Bass Coast", "Baw Baw", "East Gippsland", "Latrobe", "South Gippsland", "Wellington"],
    "Goulburn": ["Greater Shepparton", "Moira", "Murrindindi", "Strathbogie"],
    "Goulburn (part)": ["Mitchell"],
    "Great South Coast": ["Corangamite", "Glenelg", "Moyne", "Southern Grampians", "Warrnambool"],
    "Loddon Campaspe": ["Campaspe", "Central Goldfields", "Greater Bendigo", "Loddon", "Macedon Ranges", "Mount Alexander"],
    "Mallee": ["Buloke", "Gannawarra", "Mildura", "Swan Hill"],
    "Ovens Murray": ["Alpine", "Benalla", "Indigo", "Mansfield", "Towong", "Wangaratta", "Wodonga"],
    "Wimmera Southern Mallee": ["Hindmarsh", "Horsham", "Northern Grampians", "West Wimmera", "Yarriambiack"],
}


def normalise_key(value: Any) -> str:
    return _KEY_RE.sub(" ", str(value or "").upper()).strip()


def council_join_key(value: Any) -> str:
    text = normalise_key(value)
    text = text.replace("MERRI BEK", "MERRI BEK").replace("MORELAND", "MERRI BEK")
    for prefix in ("RURAL CITY OF ", "CITY OF ", "SHIRE OF ", "BOROUGH OF "):
        if text.startswith(prefix):
            text = text.removeprefix(prefix)
    text = _COUNCIL_TERMS_RE.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def source_metadata(path: Path, source_url: str, notes: str | None = None) -> dict[str, Any]:
    return {
        "name": path.name,
        "path": str(path.relative_to(ROOT)) if path.exists() else str(path),
        "source_url": source_url,
        "exists": path.exists(),
        "bytes": path.stat().st_size if path.exists() else None,
        "modified": datetime.fromtimestamp(path.stat().st_mtime, timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
        if path.exists()
        else None,
        "notes": notes,
    }


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


class _CouncilLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "a":
            self._href = dict(attrs).get("href")
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "a" and self._href is not None:
            text = " ".join(" ".join(self._text).split())
            self.links.append((text, self._href))
            self._href = None
            self._text = []


def fetch_text(url: str) -> str:
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    return urlopen(request, timeout=60).read().decode("utf-8", errors="replace")


def strip_html(html: str) -> str:
    html = re.sub(r"<script\b.*?</script>", " ", html, flags=re.S | re.I)
    html = re.sub(r"<style\b.*?</style>", " ", html, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", text).strip()


def parse_vec_structure(snippet: str) -> dict[str, Any]:
    lower = snippet.lower()
    patterns = [
        (r"has\s+(\d+)\s+wards?,?\s+with\s+one\s+councillor\s+in\s+each\s+ward", "single-councillor wards"),
        (r"has\s+(\d+)\s+wards?,?\s+with\s+(\d+)\s+councillors?\s+in\s+each\s+ward", "multi-councillor wards"),
        (r"has\s+(\d+)\s+wards?\s+with\s+(\d+)\s+councillors?\s+each", "multi-councillor wards"),
        (r"has\s+(\d+)\s+wards?\s+with\s+(\d+)\s+councillors?\s+\((\d+)\s+councillors?\s+in\s+each\s+ward\)", "multi-councillor wards"),
    ]
    for pattern, structure in patterns:
        match = re.search(pattern, lower)
        if not match:
            continue
        ward_count = int(match.group(1))
        if structure == "single-councillor wards":
            councillor_count = ward_count
        elif len(match.groups()) >= 3:
            councillor_count = int(match.group(2))
        else:
            councillor_count = ward_count * int(match.group(2))
        return {
            "electoral_structure": structure,
            "ward_count": ward_count,
            "councillor_count": councillor_count,
        }

    unsubdivided_patterns = [
        r"is\s+unsubdivided,?\s+with\s+(\d+)\s+councillors?",
        r"is\s+an\s+unsubdivided\s+council\s+with\s+(\d+)\s+councillors?",
        r"has\s+(\d+)\s+councillors?\s+and\s+is\s+unsubdivided",
    ]
    for pattern in unsubdivided_patterns:
        match = re.search(pattern, lower)
        if match:
            return {
                "electoral_structure": "unsubdivided",
                "ward_count": 0,
                "councillor_count": int(match.group(1)),
            }
    return {}


def refresh_vec_cache() -> dict[str, Any]:
    html = fetch_text(VEC_SOURCE_URL)
    parser = _CouncilLinkParser()
    parser.feed(html)
    seen: set[str] = set()
    links: list[tuple[str, str]] = []
    for text, href in parser.links:
        full_url = urljoin(VEC_SOURCE_URL, href)
        if "/electoral-boundaries/local-councils/" not in full_url:
            continue
        if "Council" not in text or text in seen:
            continue
        seen.add(text)
        links.append((text, full_url))

    rows = []
    for name, url in links:
        page_text = strip_html(fetch_text(url))
        marker = re.search(rf"{re.escape(name)}\s+(?:is|has)\s+", page_text)
        if marker is None:
            marker = re.search(
                r"[A-Z][A-Za-z-]+(?:\s+[A-Z][A-Za-z-]+){0,5}\s+Council\s+(?:is\s+unsubdivided|has\s+\d+\s+wards?)",
                page_text,
            )
        start = marker.start() if marker else -1
        snippet = page_text[start:start + 900] if start >= 0 else page_text[:900]
        rows.append({
            "council": name,
            "source_url": url,
            "snippet": snippet,
            **parse_vec_structure(snippet),
        })

    payload = {
        "source_url": VEC_SOURCE_URL,
        "fetched_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "rows": rows,
    }
    VEC_CACHE_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    return payload


def latest_sheet(workbook: Any, prefix: str) -> str:
    candidates = [
        name for name in workbook.sheetnames
        if name.startswith(prefix) and re.search(r"\d{2,4}-\d{2}$", name)
    ]
    return sorted(candidates, key=lambda name: re.search(r"(\d{2,4})-(\d{2})$", name).groups())[-1]


def load_abs_asgs() -> dict[str, dict[str, Any]]:
    payload = load_json(ABS_ASGS_PATH)
    by_code: dict[str, dict[str, Any]] = {}
    by_key: dict[str, dict[str, Any]] = {}
    for feature in payload.get("features", []):
        attributes = feature.get("attributes", {})
        if attributes.get("state_name_2021") != "Victoria":
            continue
        code = str(attributes.get("lga_code_2025") or "")
        key = council_join_key(attributes.get("lga_name_2025"))
        if code:
            by_code[code] = attributes
        if key:
            by_key[key] = attributes
    return {"by_code": by_code, "by_key": by_key}


def load_lgprf() -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], str]:
    workbook = openpyxl.load_workbook(LGPRF_PATH, read_only=True, data_only=True)
    sheet_name = latest_sheet(workbook, "Indicators ")
    year = sheet_name.removeprefix("Indicators ").strip()
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    by_key: dict[str, dict[str, Any]] = defaultdict(lambda: {"indicators": {}})
    long_rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        council = row[index["Council"]]
        indicator_id = row[index["ID"]]
        if not council or not indicator_id:
            continue
        key = council_join_key(council)
        item = {
            "year": year,
            "council_key": key,
            "council": council,
            "lgprf_group": row[index["Group"]],
            "indicator_id": indicator_id,
            "description": (row[index["Description"]] or "").strip(),
            "service_provided": row[index["Service Provided"]],
            "data_applicable": row[index["Data applicable?"]],
            "result": clean_value(row[index["Result"]]),
            "target": clean_value(row[index["Target"]]) if "Target" in index else None,
            "variation": row[index["Variation"]] if "Variation" in index else None,
        }
        by_key[key]["group"] = item["lgprf_group"]
        by_key[key]["indicators"][indicator_id] = item
        long_rows.append(item)
    return dict(by_key), long_rows, year


def load_governance() -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], str]:
    workbook = openpyxl.load_workbook(GOVERNANCE_PATH, read_only=True, data_only=True)
    sheet_name = latest_sheet(workbook, "All Councils - ")
    year = sheet_name.removeprefix("All Councils - ").strip()
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    group_column = "VLGC Group" if "VLGC Group" in index else "VLGC Groups"
    by_key: dict[str, dict[str, Any]] = defaultdict(lambda: {"item_count": 0, "yes_count": 0, "no_count": 0})
    long_rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        council = row[index["Council"]]
        if not council:
            continue
        key = council_join_key(council)
        yes_no = str(row[index["YES or NO"]] or "").strip().upper()
        by_key[key]["group"] = row[index[group_column]]
        by_key[key]["item_count"] += 1
        if yes_no == "YES":
            by_key[key]["yes_count"] += 1
        elif yes_no == "NO":
            by_key[key]["no_count"] += 1
        long_rows.append({
            "year": row[index["Year"]],
            "council_key": key,
            "council": council,
            "vlgc_group": row[index[group_column]],
            "item_id": row[index["Item ID"]],
            "governance_item": row[index["Governance and Management Item"]],
            "yes_no": yes_no,
            "date_single": clean_value(row[index["Date if YES (single item/date)"]]),
            "date_multiple": clean_value(row[index["Date if YES (multiple items/dates)"]]),
            "reason_if_no": row[index["Reason(s) if answer is No"]],
        })
    for item in by_key.values():
        count = item["item_count"]
        item["yes_pct"] = round(item["yes_count"] / count, 4) if count else None
    return dict(by_key), long_rows, year


def load_vgccc() -> dict[str, dict[str, Any]]:
    workbook = openpyxl.load_workbook(VGCCC_PATH, read_only=True, data_only=True)
    sheet = workbook["Detail Data 2023-24"]
    by_key: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=13, values_only=True):
        name = row[0]
        if not name or not row[1] or not row[2]:
            continue
        key = council_join_key(name)
        by_key[key] = {
            "year": "2023-24",
            "lga_name": name,
            "lga_code": row[1],
            "region": row[2],
            "net_expenditure": row[3],
            "seifa_dis_score": row[4],
            "seifa_dis_rank_state": row[5],
            "seifa_dis_rank_country": row[6],
            "seifa_dis_rank_metro": row[7],
            "seifa_advdis_score": row[8],
            "seifa_advdis_rank_state": row[9],
            "seifa_advdis_rank_country": row[10],
            "seifa_advdis_rank_metro": row[11],
            "adult_population_2024": row[12],
            "adults_per_venue_2024": row[13],
            "egms_per_1000_adults_2024": row[14],
            "expenditure_per_adult_2024": row[15],
            "workforce_june_2024": row[16],
            "unemployed_june_2024": row[17],
            "unemployment_rate_june_2024": row[18],
        }
    return by_key


def keyed_region_map(groups: dict[str, list[str]]) -> dict[str, str]:
    return {council_join_key(name): group for group, names in groups.items() for name in names}


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build_council_master(refresh_vec: bool = False) -> dict[str, Any]:
    geography = load_json(GEOGRAPHY_PATH)
    abs_asgs = load_abs_asgs()
    lgprf, indicator_rows, lgprf_year = load_lgprf()
    governance, governance_rows, governance_year = load_governance()
    vgccc = load_vgccc()
    vec_payload = refresh_vec_cache() if refresh_vec or not VEC_CACHE_PATH.exists() else load_json(VEC_CACHE_PATH)
    vec = {council_join_key(row.get("council")): row for row in vec_payload.get("rows", [])}

    metropolitan_regions = keyed_region_map(VIF_METROPOLITAN_REGIONS)
    regional_partnerships = keyed_region_map(VIF_REGIONAL_PARTNERSHIPS)

    rows: list[dict[str, Any]] = []
    for council in geography["councils"]:
        key = council["spatial_key"]
        join_key = council_join_key(council["short_name"])
        office = council.get("office") or {}
        abs_row = abs_asgs["by_code"].get(str(council.get("abs_lga_code") or ""))
        abs_join_method = "code" if abs_row else None
        if not abs_row:
            abs_row = abs_asgs["by_key"].get(join_key)
            abs_join_method = "name" if abs_row else None
        lgprf_row = lgprf.get(join_key, {})
        governance_row = governance.get(join_key, {})
        vgccc_row = vgccc.get(join_key, {})
        vec_row = vec.get(join_key, {})
        row = {
            "council_key": key,
            "short_name": council.get("short_name"),
            "long_name": council.get("long_name"),
            "status": council.get("status"),
            "is_active": council.get("status") == "active",
            "council_category": council.get("council_category"),
            "council_type": council.get("council_type"),
            "official_name": council.get("official_name"),
            "spatial_name": council.get("spatial_name"),
            "spatial_key": key,
            "map_join_key": key,
            "lga_code": council.get("lga_code"),
            "abs_lga_code": council.get("abs_lga_code"),
            "abs_lga_code_2025": abs_row.get("lga_code_2025") if abs_row else None,
            "abs_lga_name_2025": abs_row.get("lga_name_2025") if abs_row else None,
            "abs_area_albers_sqkm": abs_row.get("area_albers_sqkm") if abs_row else None,
            "abs_shape_area": abs_row.get("st_area(shape)") if abs_row else None,
            "abs_shape_length": abs_row.get("st_length(shape)") if abs_row else None,
            "abs_join_method": abs_join_method,
            "state": council.get("state"),
            "polygon_record_count": council.get("polygon_record_count"),
            "office_township": office.get("seat_township"),
            "office_address": office.get("address"),
            "office_lat": office.get("lat"),
            "office_lon": office.get("lon"),
            "office_geocoded": "yes" if office.get("lat") is not None and office.get("lon") is not None else "missing",
            "vif_metropolitan_region": metropolitan_regions.get(join_key),
            "vif_regional_partnership": regional_partnerships.get(join_key),
            "vif_region_note": "Mitchell Shire is split between the Northern metropolitan region and Goulburn Regional Partnership in VIF definitions."
            if join_key == "MITCHELL"
            else None,
            "vec_electoral_structure": vec_row.get("electoral_structure"),
            "vec_ward_count": vec_row.get("ward_count"),
            "vec_councillor_count": vec_row.get("councillor_count"),
            "vec_source_url": vec_row.get("source_url"),
            "lgprf_latest_year": lgprf_year,
            "lgprf_group": lgprf_row.get("group"),
            "governance_latest_year": governance_year,
            "governance_item_count": governance_row.get("item_count"),
            "governance_yes_count": governance_row.get("yes_count"),
            "governance_no_count": governance_row.get("no_count"),
            "governance_yes_pct": governance_row.get("yes_pct"),
            "vgccc_year": vgccc_row.get("year"),
            "vgccc_region": vgccc_row.get("region"),
            "vgccc_lga_code": vgccc_row.get("lga_code"),
            "vgccc_net_expenditure": vgccc_row.get("net_expenditure"),
            "vgccc_seifa_dis_score": vgccc_row.get("seifa_dis_score"),
            "vgccc_seifa_dis_rank_state": vgccc_row.get("seifa_dis_rank_state"),
            "vgccc_seifa_dis_rank_country": vgccc_row.get("seifa_dis_rank_country"),
            "vgccc_seifa_dis_rank_metro": vgccc_row.get("seifa_dis_rank_metro"),
            "vgccc_seifa_advdis_score": vgccc_row.get("seifa_advdis_score"),
            "vgccc_seifa_advdis_rank_state": vgccc_row.get("seifa_advdis_rank_state"),
            "vgccc_seifa_advdis_rank_country": vgccc_row.get("seifa_advdis_rank_country"),
            "vgccc_seifa_advdis_rank_metro": vgccc_row.get("seifa_advdis_rank_metro"),
            "vgccc_adult_population_2024": vgccc_row.get("adult_population_2024"),
            "vgccc_adults_per_venue_2024": vgccc_row.get("adults_per_venue_2024"),
            "vgccc_egms_per_1000_adults_2024": vgccc_row.get("egms_per_1000_adults_2024"),
            "vgccc_expenditure_per_adult_2024": vgccc_row.get("expenditure_per_adult_2024"),
            "vgccc_workforce_june_2024": vgccc_row.get("workforce_june_2024"),
            "vgccc_unemployed_june_2024": vgccc_row.get("unemployed_june_2024"),
            "vgccc_unemployment_rate_june_2024": vgccc_row.get("unemployment_rate_june_2024"),
            "has_abs_asgs": bool(abs_row),
            "has_lgprf": bool(lgprf_row),
            "has_governance": bool(governance_row),
            "has_vec": bool(vec_row.get("councillor_count") is not None),
            "has_vgccc": bool(vgccc_row),
        }
        indicators = lgprf_row.get("indicators", {})
        for indicator_id, column in MASTER_LGPRF_FIELDS.items():
            row[column] = indicators.get(indicator_id, {}).get("result")
        rows.append(row)

    rows.sort(key=lambda item: item["short_name"])
    payload = {
        "set_id": "victorian_council_master",
        "label": "Victorian Council Master",
        "description": "Mother reference table for Victoria's 79 councils, joining canonical names, spatial identity, cohort geography, electoral structure, performance context and governance coverage.",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sources": {
            "council_geography": source_metadata(
                GEOGRAPHY_PATH,
                VICMAP_SOURCE_URL,
                "Local Vicmap-derived geography and council office coordinates built by scripts/build_council_geography.py.",
            ),
            "abs_asgs_2025": source_metadata(ABS_ASGS_PATH, ABS_ASGS_SOURCE_URL),
            "know_your_council": source_metadata(LGPRF_PATH, KYC_SOURCE_URL),
            "lgprf_performance_reporting": {"source_url": LGPRF_SOURCE_URL, "data_vic_url": DATAVIC_LGPRF_URL},
            "governance_checklist": source_metadata(GOVERNANCE_PATH, LGPRF_SOURCE_URL),
            "victoria_in_future_regions": {"source_url": VIF_SOURCE_URL, "notes": "Used for metropolitan region and Regional Partnership mappings; numeric VIF spreadsheets are not copied into the repo."},
            "vgccc_context": source_metadata(VGCCC_PATH, VGCCC_SOURCE_URL),
            "vec_electoral_structure": source_metadata(VEC_CACHE_PATH, VEC_SOURCE_URL),
        },
        "summary": {
            "councils": len(rows),
            "columns": len(MASTER_COLUMNS),
            "categories": dict(sorted(Counter(row.get("council_category") or "unknown" for row in rows).items())),
            "council_types": dict(sorted(Counter(row.get("council_type") or "unknown" for row in rows).items())),
            "metropolitan_regions": dict(sorted(Counter(row.get("vif_metropolitan_region") or "not_applicable" for row in rows).items())),
            "regional_partnerships": dict(sorted(Counter(row.get("vif_regional_partnership") or "not_applicable" for row in rows).items())),
            "coverage": {
                "abs_asgs": sum(1 for row in rows if row["has_abs_asgs"]),
                "lgprf": sum(1 for row in rows if row["has_lgprf"]),
                "governance": sum(1 for row in rows if row["has_governance"]),
                "vec": sum(1 for row in rows if row["has_vec"]),
                "vgccc": sum(1 for row in rows if row["has_vgccc"]),
                "office_points": sum(1 for row in rows if row["office_geocoded"] == "yes"),
            },
            "companion_tables": {
                "lgprf_indicators_2024_25": str(INDICATORS_CSV_PATH.relative_to(ROOT)),
                "governance_checklist_2024_25": str(GOVERNANCE_CSV_PATH.relative_to(ROOT)),
            },
        },
        "rows": rows,
        "lookup": {row["council_key"]: row for row in rows},
    }

    MASTER_JSON_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    STATIC_MASTER_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATIC_MASTER_JSON_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    write_csv(MASTER_CSV_PATH, rows, MASTER_COLUMNS)
    write_csv(INDICATORS_CSV_PATH, indicator_rows, [
        "year",
        "council_key",
        "council",
        "lgprf_group",
        "indicator_id",
        "description",
        "service_provided",
        "data_applicable",
        "result",
        "target",
        "variation",
    ])
    write_csv(GOVERNANCE_CSV_PATH, governance_rows, [
        "year",
        "council_key",
        "council",
        "vlgc_group",
        "item_id",
        "governance_item",
        "yes_no",
        "date_single",
        "date_multiple",
        "reason_if_no",
    ])
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the Victorian council master reference table.")
    parser.add_argument("--refresh-vec", action="store_true", help="Refresh cached VEC council electoral structure pages.")
    args = parser.parse_args()
    payload = build_council_master(refresh_vec=args.refresh_vec)
    print(json.dumps({
        "rows": payload["summary"]["councils"],
        "columns": payload["summary"]["columns"],
        "coverage": payload["summary"]["coverage"],
        "output": str(MASTER_JSON_PATH.relative_to(ROOT)),
    }, indent=2))


if __name__ == "__main__":
    main()
